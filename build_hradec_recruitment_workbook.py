"""
FC Hradec Králové — Jamestown Recruitment Workbook  (v3 — extensive)

Sheets:
  README            — Methodology, column guide, how to use
  Priority List     — All CLEAR UPGRADE players, ranked by Lamberts Index
  Elite Picks       — LI ≥ 30 (ELITE VALUE tier), compound signal analysis
  Lamberts Analysis    — LI statistics, top 5 per position, distributions
  All Targets       — All 551 candidates, every metric, full filters
  GK/CB/FB/DM/CM/W/FW — Per-position deep-dive tabs
  Expiring 2026     — Players with contract expiry this summer
  Budget Planner    — Cost-optimised squad build within €1M
  Squad             — Current Hradec squad with Impect quality scores
"""

import os, warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import (
    ColorScaleRule, CellIsRule, FormulaRule, DataBarRule, IconSetRule
)
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabel

warnings.filterwarnings("ignore")

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DATA_FILE  = os.path.join(BASE_DIR, "hradec_recruitment_2526.xlsx")
SQUAD_FILE = os.path.join(BASE_DIR, "hradec_player_tracking.xlsx")
OUT_FILE   = os.path.join(BASE_DIR, "hradec_recruitment_model.xlsx")

POSITION_ORDER  = ["GK", "CB", "FB", "DM", "CM", "W", "FW"]
POSITION_LABELS = {
    "GK": "Goalkeeper", "CB": "Centre-Back", "FB": "Full-Back",
    "DM": "Defensive Mid", "CM": "Central Mid", "W": "Winger", "FW": "Forward",
}
IMPECT_MAP = {
    "GOALKEEPER": "GK", "CENTRAL_DEFENDER": "CB",
    "RIGHT_CENTRAL_DEFENDER": "CB", "LEFT_CENTRAL_DEFENDER": "CB",
    "RIGHT_DEFENDER": "FB", "LEFT_DEFENDER": "FB",
    "RIGHT_WINGBACK_DEFENDER": "FB", "LEFT_WINGBACK_DEFENDER": "FB",
    "DEFENSE_MIDFIELD": "DM", "CENTRAL_MIDFIELD": "CM",
    "ATTACKING_MIDFIELD": "CM", "RIGHT_WINGER": "W",
    "LEFT_WINGER": "W", "CENTER_FORWARD": "FW", "SECOND_STRIKER": "FW",
}

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY       = "1E3A8A"
BLUE       = "1A56DB"
BLUE_LIGHT = "DBEAFE"
BLUE_MID   = "93C5FD"
GREEN_D    = "14532D"
GREEN_M    = "166534"
GREEN_L    = "D1FAE5"
AMBER_D    = "78350F"
AMBER_M    = "92400E"
AMBER_L    = "FEF3C7"
RED_D      = "7F1D1D"
RED_M      = "991B1B"
RED_L      = "FEE2E2"
GOLD       = "F59E0B"
GOLD_L     = "FFF9C4"
PURPLE_D   = "4C1D95"
PURPLE_L   = "EDE9FE"
GREY_BG    = "F3F4F6"
GREY_MID   = "9CA3AF"
GREY_D     = "374151"
WHITE      = "FFFFFF"
BLACK      = "111111"
STRIPE     = "F8FAFF"

def fill(hex_col): return PatternFill("solid", fgColor=hex_col)
def font(bold=False, color=BLACK, size=10, italic=False, underline=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                underline="single" if underline else None, name="Calibri")
def align(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
def border_bottom(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(bottom=s)
def border_all(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom(color="9CA3AF"):
    t = Side(style="medium", color=color)
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=t)

# ── Main column definitions ──────────────────────────────────────────────────
MAIN_COLS = [
    ("Player",                      "Player",           22),
    ("Team",                        "Team",             20),
    ("league",                      "League",           13),
    ("pos_group",                   "Pos",               6),
    ("Position",                    "Full Position",    13),
    ("Age",                         "Age",               6),
    ("Contract expires",            "Contract",         11),
    ("contract_flag",               "Exp?",              6),
    ("Market value",                "Mkt Val (€)",      12),
    ("model_value",                 "Model Val (€)",    13),
    ("value_ratio",                 "Val Ratio",        10),
    ("value_tier",                  "Tier",             12),
    ("sqs_rank",                    "SQS Rank",         10),
    ("bloom_index",                 "Lamberts Index",      12),
    ("upgrade_flag",                "Status",           16),
    ("vs_hradec_gap",               "vs Hradec",        11),
    ("Minutes played",              "Minutes",          10),
    ("Goals per 90",                "Goals/90",          9),
    ("xG per 90",                   "xG/90",             8),
    ("Assists per 90",              "Assists/90",        9),
    ("xA per 90",                   "xA/90",             8),
    ("Progressive passes per 90",   "Prog Pass/90",     12),
    ("Progressive runs per 90",     "Prog Run/90",      11),
    ("Touches in box per 90",       "Box Touch/90",     12),
    ("Dribbles per 90",             "Dribbles/90",      11),
    ("Successful dribbles, %",      "Drib Succ %",      11),
    ("Defensive duels won, %",      "Def Duel %",       10),
    ("Aerial duels won, %",         "Aerial %",          9),
    ("PAdj Interceptions",          "Interceptions",    13),
    ("Key passes per 90",           "Key Pass/90",      11),
    ("Save rate, %",                "Save %",            8),
    ("Prevented goals per 90",      "Prev Goals/90",    13),
]

STATUS_FILLS = {
    "CLEAR UPGRADE":      (GREEN_L,  GREEN_M),
    "ROTATIONAL / COVER": (AMBER_L,  AMBER_M),
    "DEPTH":              (GREY_BG,  GREY_MID),
}
TIER_FILLS = {
    "ELITE VALUE":  (GREEN_L,   GREEN_M),
    "HIGH VALUE":   ("DCFCE7",  "166534"),
    "VALUE":        (BLUE_LIGHT, BLUE),
    "FAIR PRICE":   (GREY_BG,   GREY_MID),
    "OVERVALUED":   (RED_L,     RED_M),
}

# ── Helper: style a header row ───────────────────────────────────────────────
def style_header(ws, row_num, bg=NAVY, fg=WHITE, size=10, height=20):
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill      = fill(bg)
            cell.font      = font(bold=True, color=fg, size=size)
            cell.alignment = align("center")
            cell.border    = border_all("334155")
    ws.row_dimensions[row_num].height = height

def style_subheader(ws, row_num, bg=BLUE_LIGHT, fg=BLUE, size=9):
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill      = fill(bg)
            cell.font      = font(color=fg, size=size)
            cell.alignment = align("left", "center")

def title_row(ws, row_num, text, cols, bg=NAVY, fg=WHITE, size=13, height=28):
    ws.cell(row_num, 1, text)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=cols)
    c = ws.cell(row_num, 1)
    c.fill = fill(bg); c.font = font(bold=True, color=fg, size=size)
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[row_num].height = height

def subtitle_row(ws, row_num, text, cols, bg=BLUE_LIGHT, fg=BLUE, size=9, height=14):
    ws.cell(row_num, 1, text)
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=cols)
    c = ws.cell(row_num, 1)
    c.fill = fill(bg); c.font = font(color=fg, size=size, italic=True)
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[row_num].height = height

# ── Build DataFrame from raw data ────────────────────────────────────────────
def build_df(df_raw: pd.DataFrame, cols=None) -> pd.DataFrame:
    if cols is None: cols = MAIN_COLS
    src = [c[0] for c in cols]; hdr = [c[1] for c in cols]
    existing = [c for c in src if c in df_raw.columns]
    out = df_raw[existing].copy()
    rename = {c[0]: c[1] for c in cols if c[0] in out.columns}
    out = out.rename(columns=rename)
    for h in hdr:
        if h not in out.columns: out[h] = ""
    return out[hdr]

# ── Write a full data table to a worksheet ──────────────────────────────────
def write_table(ws, df, title_text, subtitle_text="",
                start_row=1, table_name=None, table_style="TableStyleMedium2"):
    ncols = len(df.columns)
    title_row(ws, start_row, title_text, ncols)
    if subtitle_text:
        subtitle_row(ws, start_row+1, subtitle_text, ncols)
        hrow = start_row + 2
    else:
        hrow = start_row + 1

    ws.append(list(df.columns))
    style_header(ws, hrow)

    col_idx_map = {col: i+1 for i, col in enumerate(df.columns)}

    for r_offset, row in enumerate(df.itertuples(index=False)):
        r = hrow + 1 + r_offset
        ws.append(list(row))
        ws.row_dimensions[r].height = 15
        bg = STRIPE if r_offset % 2 == 0 else WHITE
        for cell in ws[r]:
            cell.fill = fill(bg)
            cell.font = font(size=9)
            cell.alignment = align("left", "center")
            cell.border = border_bottom()

        # ── Status colour
        if "Status" in col_idx_map:
            ci = col_idx_map["Status"]
            v  = ws.cell(r, ci).value
            if v in STATUS_FILLS:
                bg2, fg2 = STATUS_FILLS[v]
                c = ws.cell(r, ci)
                c.fill = fill(bg2); c.font = font(bold=True, color=fg2, size=9)
                c.alignment = align("center", "center")

        # ── Tier colour
        if "Tier" in col_idx_map:
            ci = col_idx_map["Tier"]
            v  = ws.cell(r, ci).value
            if v in TIER_FILLS:
                bg2, fg2 = TIER_FILLS[v]
                c = ws.cell(r, ci)
                c.fill = fill(bg2); c.font = font(bold=True, color=fg2, size=9)
                c.alignment = align("center", "center")

        # ── SQS Rank colour
        if "SQS Rank" in col_idx_map:
            ci = col_idx_map["SQS Rank"]
            try:
                v = float(ws.cell(r, ci).value or 0)
                col = GREEN_M if v >= 70 else AMBER_M if v >= 40 else RED_M
                ws.cell(r, ci).font = font(bold=True, color=col, size=9)
                ws.cell(r, ci).alignment = align("center", "center")
            except: pass

        # ── Lamberts Index colour + bold
        if "Lamberts Index" in col_idx_map:
            ci = col_idx_map["Lamberts Index"]
            try:
                v = float(ws.cell(r, ci).value or 0)
                col = (GREEN_M if v >= 30 else GREEN_M if v >= 20 else
                       BLUE if v >= 10 else RED_M if v < 0 else GREY_D)
                ws.cell(r, ci).font = font(bold=True, color=col, size=10)
                ws.cell(r, ci).alignment = align("center", "center")
            except: pass

        # ── vs Hradec colour
        if "vs Hradec" in col_idx_map:
            ci = col_idx_map["vs Hradec"]
            try:
                v = float(ws.cell(r, ci).value or 0)
                col = GREEN_M if v > 5 else RED_M if v < -5 else GREY_D
                ws.cell(r, ci).font = font(bold=True, color=col, size=9)
                ws.cell(r, ci).alignment = align("center", "center")
            except: pass

        # ── Age colour
        if "Age" in col_idx_map:
            ci = col_idx_map["Age"]
            try:
                v = int(ws.cell(r, ci).value or 0)
                col = GREEN_M if v <= 23 else GREY_D if v <= 27 else GREY_MID
                ws.cell(r, ci).font = font(bold=(v <= 23), color=col, size=9)
                ws.cell(r, ci).alignment = align("center", "center")
            except: pass

        # ── Expiry flag
        if "Exp?" in col_idx_map:
            ci = col_idx_map["Exp?"]
            v  = ws.cell(r, ci).value
            if str(v) in ("2026", "2027"):
                ws.cell(r, ci).font = font(bold=True, color=RED_M if str(v) == "2026" else AMBER_M, size=9)
                ws.cell(r, ci).alignment = align("center", "center")

        # ── Val Ratio colour
        if "Val Ratio" in col_idx_map:
            ci = col_idx_map["Val Ratio"]
            try:
                v = float(str(ws.cell(r, ci).value or "0").replace("×",""))
                col = GREEN_M if v >= 2 else BLUE if v >= 1.5 else GREY_D if v >= 1 else RED_M
                ws.cell(r, ci).font = font(bold=(v >= 1.5), color=col, size=9)
                ws.cell(r, ci).alignment = align("center", "center")
            except: pass

        # ── Market value number format
        for col_name in ("Mkt Val (€)", "Model Val (€)"):
            if col_name in col_idx_map:
                ci = col_idx_map[col_name]
                try:
                    v = float(ws.cell(r, ci).value or 0)
                    ws.cell(r, ci).number_format = '€#,##0'
                except: pass

    # Column widths
    width_map = {c[1]: c[2] for c in MAIN_COLS}
    for col_i, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = width_map.get(col_name, 12)

    # Auto-filter on header row
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(ncols)}{hrow}"
    ws.freeze_panes = f"B{hrow+1}"

    # Lamberts Index color scale
    if "Lamberts Index" in col_idx_map:
        bl = get_column_letter(col_idx_map["Lamberts Index"])
        dr = hrow + 1; er = hrow + len(df)
        ws.conditional_formatting.add(f"{bl}{dr}:{bl}{er}", ColorScaleRule(
            start_type="num", start_value=-50, start_color="C0392B",
            mid_type="num",   mid_value=0,     mid_color="FFFFFF",
            end_type="num",   end_value=50,     end_color="1E8449",
        ))

    # SQS Rank data bar
    if "SQS Rank" in col_idx_map:
        sl = get_column_letter(col_idx_map["SQS Rank"])
        dr = hrow + 1; er = hrow + len(df)
        ws.conditional_formatting.add(f"{sl}{dr}:{sl}{er}", DataBarRule(
            start_type="num", start_value=0, end_type="num", end_value=100,
            color="1A56DB", showValue=True
        ))

    ws.sheet_view.showGridLines = False
    return hrow


# ══════════════════════════════════════════════════════════════════════════════
# README SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_readme(ws):
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 75
    ws.column_dimensions["D"].width = 2

    def w(row, col, val, bold=False, size=10, color=BLACK, bg=None,
          merge=None, wrap=False, italic=False, align_h="left"):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font(bold=bold, size=size, color=color, italic=italic)
        c.alignment = align(align_h, "center", wrap=wrap)
        if bg: c.fill = fill(bg)
        if merge: ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge)
        return c

    title_row(ws, 1, "FC HRADEC KRÁLOVÉ — JAMESTOWN RECRUITMENT MODEL  2025–26", 20,
              bg=NAVY, size=14, height=32)
    subtitle_row(ws, 2,
        "Waltzing Analytics  ·  Jamestown / Marc Lamberts methodology  ·  CZ II + Slovakia + Slovakia II  ·  Budget ≤ €1,000,000  ·  Age ≤ 30",
        20, height=15)

    # Section headers helper
    def sec(row, label):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=20)
        c = ws.cell(row, 2, label)
        c.fill = fill(NAVY); c.font = font(bold=True, color=WHITE, size=11)
        c.alignment = align("left", "center", indent=1)
        ws.row_dimensions[row].height = 22

    def row2(row, b_val, c_val, b_bold=False, c_wrap=True, h=None):
        w(row, 2, b_val, bold=b_bold, size=9, color=GREY_D, bg=GREY_BG if b_bold else WHITE)
        w(row, 3, c_val, size=9, color=BLACK, wrap=c_wrap)
        ws.row_dimensions[row].height = h or (30 if c_wrap and len(str(c_val)) > 60 else 15)

    # ── WORKBOOK STRUCTURE
    sec(4, "WORKBOOK STRUCTURE")
    row2(5,  "Sheet", "Contents", b_bold=True, h=16)
    row2(6,  "README",           "This sheet — column guide, methodology, how to use.")
    row2(7,  "Priority List",    "ALL clear upgrade players across every position, ranked by Lamberts Index. Start here for recruitment priorities.")
    row2(8,  "Elite Picks",      "Only ELITE VALUE tier (LI ≥ 30). Compound signal table — players where LI ≥ 30 AND model undervalues AND clear upgrade. Highest-conviction targets.")
    row2(9,  "Lamberts Analysis",   "Lamberts Index statistics summary — average, max, % positive per position. Top 5 players by position. LI distribution table. Compound signals.")
    row2(10, "All Targets",      "All 551 candidates with every metric and full column filters. Ranked by Lamberts Index by default.")
    row2(11, "GK / CB / FB…",   "One tab per position. Pre-filtered to that role only, sorted by Lamberts Index descending.")
    row2(12, "Expiring 2026",    "Players whose contracts expire in 2026. Many can be acquired for free or at a heavy discount — especially valuable targets.")
    row2(13, "Budget Planner",   "Cost-optimised squad build. Elite Value + Clear Upgrade targets sorted by market value. Shows running total to help build within the €1M budget.")
    row2(14, "Squad",            "Current FC Hradec Králové squad quality scores from Impect API. Shows quality percentile by position — basis for the 'vs Hradec' gap calculations.")

    # ── COLUMN GUIDE
    sec(16, "COLUMN GUIDE")
    row2(17, "Column", "Meaning", b_bold=True, h=16)
    row2(18, "SQS Rank (0–100)",  "Statistical Quality Score percentile rank within the player's position group across the full recruitment pool. Built from position-weighted per-90 Wyscout metrics (Goals, xG, xA, Prog passes, Dribbles, Def duels, Aerial duels, Interceptions etc.) adjusted by league difficulty coefficient. Green ≥70, Amber ≥40, Red <40.")
    row2(19, "Lamberts Index",        "The core Jamestown signal. SQS rank minus Market Value rank within the same position group. POSITIVE = player performing above what the market prices in (undervalued). Negative = overvalued by market. Range: −100 to +100. Green ≥20, Blue ≥10, Red <0.")
    row2(20, "Tier",               "ELITE VALUE: LI ≥ 30 — strongest buy signal. HIGH VALUE: LI ≥ 20. VALUE: LI ≥ 10. FAIR PRICE: 0–9, no clear edge. OVERVALUED: LI < 0.")
    row2(21, "vs Hradec",          "SQS rank gap vs current Hradec starters at that position (top 2 by minutes, from Impect data). +25 means 25 percentile points better than who is currently playing. Green > +5, Red < −5.")
    row2(22, "Status",             "CLEAR UPGRADE: SQS rank > starter average + 10 pts — immediate first-team improvement. ROTATIONAL / COVER: near-equivalent quality, provides competition. DEPTH: below current starter level.")
    row2(23, "Model Val (€)",      "XGBoost model estimate of fair market value based on stats alone. Uses 5-fold out-of-fold predictions (never predicts on training data). OOF R² ≈ 0.12 — low R² is expected since market value is reputation-driven, not stats-driven. Gap between Model Val and Mkt Val reveals mispricing.")
    row2(24, "Val Ratio",          "Model Val ÷ Market Val. Ratio ≥ 2.0× means the model thinks the player is worth at least twice what the market charges — a strong compound signal when combined with positive Lamberts Index.")
    row2(25, "Contract / Exp?",    "Contract expiry from Wyscout. Exp? column: '2026' = expiring this summer (can often be pre-contracted or acquired cheaply). '2027' = expiring next summer (negotiation opportunity).")
    row2(26, "Age",                "Green ≤23 (development asset with sell-on value). Grey 24–27 (prime window). Light grey 28+ (short resale window — lower value in the asset model).")
    row2(27, "League",             "CZ II = Czech second tier (difficulty coefficient ×0.82). Slovakia = Slovak Superliga (×0.78). Slovakia II = Slovak second tier (×0.68). SQS already accounts for this — a CZ II stat scores higher than the same stat in Slovakia II.")

    # ── METHODOLOGY
    sec(29, "METHODOLOGY — HOW THE MODEL WORKS")
    row2(30, "Overview",           "The Jamestown / Marc Lamberts approach: in lower leagues, market values are driven by reputation and agent relationships, NOT by objective statistical output. This creates systematic and exploitable price inefficiencies. The model quantifies this gap.", h=30)
    row2(31, "Step 1 — Data",      "Load Wyscout Market I files for CZ II, Slovakia, Slovakia II (2025-26 season). Filter: ≥900 minutes played, Age ≤30, Market Value ≤€1M. 551 players pass these criteria.")
    row2(32, "Step 2 — SQS",       "Compute position-weighted per-90 stats score. Weight matrix tailored to each position (FW: Goals/xG/Box touches highest; DM: Interceptions/Def duels highest etc.). Apply league difficulty multiplier (CZ II ×0.82, Slovakia ×0.78, Slovakia II ×0.68). Percentile-rank within position group → SQS Rank 0–100.")
    row2(33, "Step 3 — Model Val", "Train XGBoost to predict log(market value) from age, position, league, minutes, SQS components. 5-fold stratified cross-validation. Out-of-fold predictions only — model never predicts on data it trained on. Exponentiate back to euros.")
    row2(34, "Step 4 — Lamberts LI",  "Lamberts Index = SQS percentile rank − Market Value percentile rank (both ranked within position group). Positive LI = underpriced. Tier thresholds: ELITE ≥30, HIGH ≥20, VALUE ≥10, FAIR 0–9, OVER <0.")
    row2(35, "Step 5 — vs Hradec", "Load Impect tracking data for FC Hradec Králové (current season). Identify top 2 players by minutes at each position. Compute their average SQS-equivalent quality percentile. vs Hradec = candidate SQS rank − Hradec starter average. CLEAR UPGRADE if gap ≥ +10.")

    # ── HOW TO USE
    sec(37, "HOW TO USE THIS WORKBOOK")
    row2(38, "Quick start",        "Priority List → sort Lamberts Index descending → filter Status = CLEAR UPGRADE → filter Age ≤ 26 for young targets. These are your highest-conviction picks.")
    row2(39, "Elite targets",      "Elite Picks tab → compound signal table at bottom. Players where LI ≥ 30 AND Val Ratio ≥ 2 AND CLEAR UPGRADE are the absolute strongest buy signals.")
    row2(40, "Budget shopping",    "Budget Planner tab → shows elite/high-value clear upgrades sorted by price with running total. Plan a multi-player recruitment window within €1M.")
    row2(41, "Expiring contracts", "Expiring 2026 tab → players whose deals run out this summer. Pre-contracts, free transfers, or significant discounts. Filter by position to find your best options.")
    row2(42, "Position deep-dive", "Click any position tab (GK/CB/FB/DM/CM/W/FW). Pre-filtered, sorted by BI. Sort by SQS Rank to see best performers regardless of price. Sort by Mkt Val for cheapest options.")
    row2(43, "Colour quick guide", "Status: Green=Upgrade · Amber=Rotational · Grey=Depth\nBloom: Deep Green≥30 · Green≥20 · Blue≥10 · Red<0\nSQS: Green≥70 · Amber≥40 · Red<40\nVal Ratio: Green≥2× · Blue≥1.5× · Red<1×", h=40)


# ══════════════════════════════════════════════════════════════════════════════
# BLOOM ANALYSIS SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_bloom_analysis(ws, df):
    ws.title = "Lamberts Analysis"
    ws.sheet_view.showGridLines = False

    ncols = 14
    title_row(ws, 1, "BLOOM INDEX ANALYSIS — COMPLETE BREAKDOWN", ncols,
              bg=NAVY, size=13, height=28)
    subtitle_row(ws, 2,
        "LI = SQS Percentile Rank − Market Value Percentile Rank  ·  ELITE ≥30  ·  HIGH ≥20  ·  VALUE ≥10  ·  FAIR 0–9  ·  OVER <0",
        ncols)

    def hdr(row, vals, bg=NAVY, fg=WHITE, h=18):
        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i, v)
            c.fill = fill(bg); c.font = font(bold=True, color=fg, size=9)
            c.alignment = align("center")
            c.border = border_all("334155")
        ws.row_dimensions[row].height = h

    def cell(r, c, v, bold=False, color=BLACK, bg=WHITE, align_h="center", fmt=None):
        cell = ws.cell(r, c, v)
        cell.fill = fill(bg); cell.font = font(bold=bold, color=color, size=9)
        cell.alignment = align(align_h, "center")
        cell.border = border_bottom()
        if fmt: cell.number_format = fmt
        return cell

    # ── SECTION 1: Overall statistics
    row = 4
    ws.cell(row, 1, "OVERALL LI STATISTICS").fill = fill(BLUE_LIGHT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1)
    c.font = font(bold=True, color=BLUE, size=10)
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    total = len(df)
    avg_bi = round(df["bloom_index"].mean(), 1)
    max_bi = round(df["bloom_index"].max(), 1)
    min_bi = round(df["bloom_index"].min(), 1)
    pct_pos = round((df["bloom_index"] > 0).sum() / total * 100, 1)
    elite_n = (df["value_tier"] == "ELITE VALUE").sum()
    high_n  = (df["value_tier"] == "HIGH VALUE").sum()
    val_n   = (df["value_tier"] == "VALUE").sum()
    upgr_n  = (df["upgrade_flag"] == "CLEAR UPGRADE").sum()
    compound = df[(df["bloom_index"] >= 20) & (df["value_ratio"] >= 2) & (df["upgrade_flag"] == "CLEAR UPGRADE")].shape[0]
    exp2026_n= df["contract_flag"].astype(str).eq("2026").sum()

    stats_labels = ["Total Candidates", "Avg Lamberts Index", "Max Lamberts Index",
                    "Min Lamberts Index", "% Positive BI", "ELITE VALUE (≥30)",
                    "HIGH VALUE (≥20)", "VALUE (≥10)", "Clear Upgrades", "Compound Signals", "Expiring 2026"]
    stats_values = [total, avg_bi, max_bi, min_bi, f"{pct_pos}%", elite_n,
                    high_n, val_n, upgr_n, compound, exp2026_n]
    stats_colors = [BLACK, BLUE if avg_bi > 0 else RED_M, GREEN_M, RED_M,
                    GREEN_M if pct_pos > 50 else AMBER_M,
                    GREEN_M, GREEN_M, BLUE, GREEN_M, PURPLE_D, AMBER_M]

    hdr(row, ["Metric", "Value", "", "Tier Breakdown", "Count", "%", "", "League", "Count", "Avg BI", "", "Compound Signals", "Count", ""])
    row += 1

    tiers  = ["ELITE VALUE", "HIGH VALUE", "VALUE", "FAIR PRICE", "OVERVALUED"]
    tier_c = [GREEN_M, GREEN_M, BLUE, GREY_MID, RED_M]
    leagues = df["league"].unique()
    for i in range(max(len(stats_labels), len(tiers))):
        bg = STRIPE if i % 2 == 0 else WHITE
        ws.row_dimensions[row].height = 15
        if i < len(stats_labels):
            cell(row, 1, stats_labels[i], bg=bg, align_h="left")
            cell(row, 2, stats_values[i], bold=True, color=stats_colors[i], bg=bg)
        if i < len(tiers):
            n = (df["value_tier"] == tiers[i]).sum()
            pct = round(n / total * 100, 1)
            cell(row, 4, tiers[i], bold=True, color=tier_c[i], bg=bg, align_h="left")
            cell(row, 5, n, bold=True, color=tier_c[i], bg=bg)
            cell(row, 6, f"{pct}%", color=tier_c[i], bg=bg)
        if i < len(leagues):
            lg = sorted(leagues)[i] if i < len(sorted(leagues)) else ""
            lgdf = df[df["league"] == lg]
            if len(lgdf):
                lg_avg = round(lgdf["bloom_index"].mean(), 1)
                cell(row, 8, lg, bg=bg, align_h="left")
                cell(row, 9, len(lgdf), bold=True, color=BLUE, bg=bg)
                cell(row, 10, lg_avg, bold=True, color=GREEN_M if lg_avg > 0 else RED_M, bg=bg)
        row += 1

    # ── SECTION 2: LI Statistics per position
    row += 1
    ws.cell(row, 1, "BLOOM INDEX BY POSITION").fill = fill(BLUE_LIGHT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1)
    c.font = font(bold=True, color=BLUE, size=10)
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    hdr(row, ["Position", "Count", "Avg BI", "Max BI", "Min BI", "% Pos BI",
              "ELITE", "HIGH", "VALUE", "Clear Upg", "Exp 2026", "Best LI Player", "Team", "BI"])
    row += 1

    for pi, pos in enumerate(POSITION_ORDER):
        pdf = df[df["pos_group"] == pos]
        if len(pdf) == 0: continue
        bg = STRIPE if pi % 2 == 0 else WHITE
        ws.row_dimensions[row].height = 15
        a_bi = round(pdf["bloom_index"].mean(), 1)
        m_bi = round(pdf["bloom_index"].max(), 1)
        n_bi = round(pdf["bloom_index"].min(), 1)
        pp   = round((pdf["bloom_index"] > 0).sum() / len(pdf) * 100, 1)
        el   = (pdf["value_tier"] == "ELITE VALUE").sum()
        hi   = (pdf["value_tier"] == "HIGH VALUE").sum()
        va   = (pdf["value_tier"] == "VALUE").sum()
        cu   = (pdf["upgrade_flag"] == "CLEAR UPGRADE").sum()
        ex   = pdf["contract_flag"].astype(str).eq("2026").sum()
        best = pdf.nlargest(1, "bloom_index").iloc[0] if len(pdf) else None

        cell(row, 1, POSITION_LABELS[pos], bold=True, color=NAVY, bg=bg, align_h="left")
        cell(row, 2, len(pdf), bg=bg)
        cell(row, 3, a_bi, bold=True, color=GREEN_M if a_bi > 0 else RED_M, bg=bg)
        cell(row, 4, m_bi, bold=True, color=GREEN_M, bg=bg)
        cell(row, 5, n_bi, bold=True, color=RED_M, bg=bg)
        cell(row, 6, f"{pp}%", bg=bg)
        cell(row, 7, el, bold=True, color=GREEN_M if el > 0 else GREY_MID, bg=bg)
        cell(row, 8, hi, bold=True, color=GREEN_M if hi > 0 else GREY_MID, bg=bg)
        cell(row, 9, va, bold=True, color=BLUE if va > 0 else GREY_MID, bg=bg)
        cell(row, 10, cu, bold=True, color=GREEN_M if cu > 0 else GREY_MID, bg=bg)
        cell(row, 11, ex, bold=True, color=AMBER_M if ex > 0 else GREY_MID, bg=bg)
        if best is not None:
            cell(row, 12, str(best["Player"]), bold=True, color=NAVY, bg=bg, align_h="left")
            cell(row, 13, str(best["Team"]), bg=bg, align_h="left")
            cell(row, 14, round(float(best["bloom_index"]), 1), bold=True, color=GREEN_M, bg=bg)
        row += 1

    # ── SECTION 3: LI Distribution
    row += 1
    ws.cell(row, 1, "BLOOM INDEX DISTRIBUTION").fill = fill(BLUE_LIGHT)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row, 1)
    c.font = font(bold=True, color=BLUE, size=10)
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    bins = [(-100,-40,"Very Undervalued"), (-40,-20,"Undervalued"), (-20,-10,"Slightly Under"),
            (-10,0,"Marginal"), (0,10,"Neutral"), (10,20,"VALUE"),
            (20,30,"HIGH VALUE"), (30,50,"ELITE VALUE"), (50,100,"Exceptional")]
    bin_colors = [RED_D, RED_M, AMBER_M, AMBER_M, GREY_D, BLUE, GREEN_M, GREEN_D, PURPLE_D]

    hdr(row, ["BI Range", "Label", "Count", "%", "Bar"])
    row += 1
    for bi2, (lo, hi2, lbl) in enumerate(bins):
        n = ((df["bloom_index"] >= lo) & (df["bloom_index"] < hi2)).sum()
        pct = round(n / total * 100, 1)
        bg = STRIPE if bi2 % 2 == 0 else WHITE
        ws.row_dimensions[row].height = 15
        cell(row, 1, f"{lo} to {hi2}", bg=bg, align_h="left")
        cell(row, 2, lbl, bold=True, color=bin_colors[bi2], bg=bg, align_h="left")
        cell(row, 3, n, bold=True, bg=bg)
        cell(row, 4, f"{pct}%", bg=bg)
        bar = "█" * max(0, int(pct / 2))
        cell(row, 5, bar, color=bin_colors[bi2], bg=bg, align_h="left")
        row += 1

    # ── SECTION 4: Compound Signals
    row += 1
    ws.cell(row, 1, "COMPOUND SIGNALS — LI ≥ 20  AND  Val Ratio ≥ 2×  AND  CLEAR UPGRADE  (Highest-conviction targets)").fill = fill(GREEN_L)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1)
    c.font = font(bold=True, color=GREEN_D, size=10)
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[row].height = 18
    row += 1

    comp_df = df[(df["bloom_index"] >= 20) & (df["value_ratio"] >= 2) & (df["upgrade_flag"] == "CLEAR UPGRADE")].sort_values("bloom_index", ascending=False)
    hdr(row, ["#", "Player", "Team", "League", "Pos", "Age", "Contract",
              "Mkt Val (€)", "Model Val (€)", "Val Ratio", "SQS Rank", "Lamberts LI", "vs Hradec", "Tier"])
    row += 1
    for ci2, (_, r2) in enumerate(comp_df.iterrows()):
        bg = GREEN_L if ci2 % 2 == 0 else "E8FEF0"
        ws.row_dimensions[row].height = 15
        ws.cell(row, 1, ci2+1).fill = fill(bg)
        ws.cell(row, 2, str(r2["Player"])).fill = fill(bg)
        ws.cell(row, 2).font = font(bold=True, color=GREEN_D, size=9)
        ws.cell(row, 2).alignment = align("left", "center")
        ws.cell(row, 3, str(r2["Team"])).fill = fill(bg); ws.cell(row, 3).font = font(size=9)
        ws.cell(row, 3).alignment = align("left", "center")
        ws.cell(row, 4, str(r2["league"])).fill = fill(bg); ws.cell(row, 4).font = font(size=9)
        ws.cell(row, 4).alignment = align("left", "center")
        ws.cell(row, 5, str(r2["pos_group"])).fill = fill(bg); ws.cell(row, 5).font = font(bold=True, color=BLUE, size=9)
        ws.cell(row, 5).alignment = align("center", "center")
        ws.cell(row, 6, int(r2["Age"] or 0)).fill = fill(bg)
        ws.cell(row, 6).font = font(bold=True, color=GREEN_M if int(r2["Age"] or 0) <= 23 else GREY_D, size=9)
        ws.cell(row, 6).alignment = align("center", "center")
        ws.cell(row, 7, str(r2.get("Contract expires",""))[:10]).fill = fill(bg); ws.cell(row, 7).font = font(size=9)
        ws.cell(row, 7).alignment = align("center", "center")
        ws.cell(row, 8, int(r2.get("Market value", 0) or 0)).fill = fill(bg)
        ws.cell(row, 8).font = font(size=9); ws.cell(row, 8).number_format = '€#,##0'
        ws.cell(row, 8).alignment = align("right", "center")
        ws.cell(row, 9, int(r2.get("model_value", 0) or 0)).fill = fill(bg)
        ws.cell(row, 9).font = font(size=9); ws.cell(row, 9).number_format = '€#,##0'
        ws.cell(row, 9).alignment = align("right", "center")
        vr = round(r2.get("value_ratio", 0) or 0, 1)
        ws.cell(row, 10, f"{vr}×").fill = fill(bg)
        ws.cell(row, 10).font = font(bold=True, color=GREEN_M, size=9); ws.cell(row, 10).alignment = align("center", "center")
        ws.cell(row, 11, round(r2.get("sqs_rank", 0) or 0, 1)).fill = fill(bg)
        ws.cell(row, 11).font = font(bold=True, color=GREEN_M, size=9); ws.cell(row, 11).alignment = align("center", "center")
        bi_v = round(r2.get("bloom_index", 0) or 0, 1)
        ws.cell(row, 12, bi_v).fill = fill(bg)
        ws.cell(row, 12).font = font(bold=True, color=GREEN_D, size=10); ws.cell(row, 12).alignment = align("center", "center")
        gap = round(r2.get("vs_hradec_gap", 0) or 0, 1)
        ws.cell(row, 13, gap).fill = fill(bg)
        ws.cell(row, 13).font = font(bold=True, color=GREEN_M, size=9); ws.cell(row, 13).alignment = align("center", "center")
        ws.cell(row, 14, str(r2.get("value_tier",""))).fill = fill(GREEN_L)
        ws.cell(row, 14).font = font(bold=True, color=GREEN_D, size=9); ws.cell(row, 14).alignment = align("center", "center")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 10
    ws.column_dimensions["M"].width = 20
    ws.column_dimensions["N"].width = 16


# ══════════════════════════════════════════════════════════════════════════════
# TOP 5 PER POSITION SHEET (part of Lamberts Analysis or separate)
# ══════════════════════════════════════════════════════════════════════════════
def build_top5_sheet(ws, df):
    ws.title = "Top 5 per Position"
    ws.sheet_view.showGridLines = False
    ncols = 13
    title_row(ws, 1, "TOP 5 PLAYERS PER POSITION — Ranked by Lamberts Index", ncols, bg=NAVY)
    subtitle_row(ws, 2, "Highest Lamberts Index player at each position group across all recruitment leagues", ncols)

    col_labels = ["#", "Player", "Team", "League", "Age", "Contract",
                  "Mkt Val (€)", "SQS Rank", "Lamberts LI", "Val Ratio", "vs Hradec", "Status", "Tier"]

    row = 3
    for pi, pos in enumerate(POSITION_ORDER):
        pdf = df[df["pos_group"] == pos].sort_values("bloom_index", ascending=False).head(5)
        if len(pdf) == 0: continue

        # Position section header
        ws.cell(row, 1, f"  {pos} — {POSITION_LABELS[pos].upper()}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row, 1)
        c.fill = fill(BLUE if pi % 2 == 0 else "1D4ED8")
        c.font = font(bold=True, color=WHITE, size=11)
        c.alignment = align("left", "center")
        ws.row_dimensions[row].height = 20
        row += 1

        # Column headers
        for ci2, lbl in enumerate(col_labels, 1):
            c = ws.cell(row, ci2, lbl)
            c.fill = fill(BLUE_LIGHT); c.font = font(bold=True, color=NAVY, size=9)
            c.alignment = align("center"); c.border = border_all("93C5FD")
        ws.row_dimensions[row].height = 16
        row += 1

        for rank, (_, r2) in enumerate(pdf.iterrows()):
            bg_map = {0: GOLD_L, 1: STRIPE, 2: STRIPE, 3: WHITE, 4: WHITE}
            bg = bg_map.get(rank, WHITE)
            ws.row_dimensions[row].height = 15

            medal = ["🥇","🥈","🥉","4.","5."][rank]
            ws.cell(row, 1, medal).fill = fill(bg)
            ws.cell(row, 1).font = font(bold=(rank<3), size=10); ws.cell(row, 1).alignment = align("center")

            ws.cell(row, 2, str(r2["Player"])).fill = fill(bg)
            ws.cell(row, 2).font = font(bold=True, color=NAVY if rank == 0 else BLACK, size=9)
            ws.cell(row, 2).alignment = align("left", "center")

            ws.cell(row, 3, str(r2["Team"])).fill = fill(bg); ws.cell(row, 3).font = font(size=9)
            ws.cell(row, 3).alignment = align("left", "center")
            ws.cell(row, 4, str(r2["league"])).fill = fill(bg); ws.cell(row, 4).font = font(size=9)
            ws.cell(row, 4).alignment = align("left", "center")

            age = int(r2.get("Age", 0) or 0)
            ws.cell(row, 5, age).fill = fill(bg)
            ws.cell(row, 5).font = font(bold=(age<=23), color=GREEN_M if age<=23 else GREY_D, size=9)
            ws.cell(row, 5).alignment = align("center")

            ws.cell(row, 6, str(r2.get("Contract expires",""))[:10]).fill = fill(bg)
            ws.cell(row, 6).font = font(size=9); ws.cell(row, 6).alignment = align("center")

            mv = int(r2.get("Market value", 0) or 0)
            ws.cell(row, 7, mv).fill = fill(bg); ws.cell(row, 7).number_format = '€#,##0'
            ws.cell(row, 7).font = font(size=9); ws.cell(row, 7).alignment = align("right")

            sqs = round(r2.get("sqs_rank", 0) or 0, 1)
            col_sqs = GREEN_M if sqs >= 70 else AMBER_M if sqs >= 40 else RED_M
            ws.cell(row, 8, sqs).fill = fill(bg); ws.cell(row, 8).font = font(bold=True, color=col_sqs, size=9)
            ws.cell(row, 8).alignment = align("center")

            bi = round(r2.get("bloom_index", 0) or 0, 1)
            col_bi = GREEN_D if bi >= 30 else GREEN_M if bi >= 20 else BLUE if bi >= 10 else RED_M
            ws.cell(row, 9, bi).fill = fill(bg); ws.cell(row, 9).font = font(bold=True, color=col_bi, size=10)
            ws.cell(row, 9).alignment = align("center")

            vr = round(r2.get("value_ratio", 0) or 0, 1)
            ws.cell(row, 10, f"{vr}×").fill = fill(bg)
            ws.cell(row, 10).font = font(bold=(vr>=2), color=GREEN_M if vr>=2 else BLUE if vr>=1 else RED_M, size=9)
            ws.cell(row, 10).alignment = align("center")

            gap = round(r2.get("vs_hradec_gap", 0) or 0, 1)
            ws.cell(row, 11, gap).fill = fill(bg)
            ws.cell(row, 11).font = font(bold=True, color=GREEN_M if gap>5 else RED_M if gap<-5 else GREY_D, size=9)
            ws.cell(row, 11).alignment = align("center")

            upgr = str(r2.get("upgrade_flag",""))
            upgr_short = "▲ Clear" if "CLEAR" in upgr else "~ Rot." if "ROT" in upgr else "Depth"
            upg_bg, upg_fg = (GREEN_L, GREEN_M) if "CLEAR" in upgr else (AMBER_L, AMBER_M) if "ROT" in upgr else (GREY_BG, GREY_MID)
            ws.cell(row, 12, upgr_short).fill = fill(upg_bg)
            ws.cell(row, 12).font = font(bold=True, color=upg_fg, size=9); ws.cell(row, 12).alignment = align("center")

            tier = str(r2.get("value_tier",""))
            t_bg, t_fg = TIER_FILLS.get(tier, (GREY_BG, GREY_MID))
            ws.cell(row, 13, tier).fill = fill(t_bg)
            ws.cell(row, 13).font = font(bold=True, color=t_fg, size=9); ws.cell(row, 13).alignment = align("center")

            row += 1

        row += 1  # gap between positions

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 12
    ws.column_dimensions["M"].width = 14
    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# EXPIRING CONTRACTS SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_expiring(ws, df):
    ws.title = "Expiring 2026"
    ws.sheet_view.showGridLines = False
    exp = df[df["contract_flag"].astype(str) == "2026"].sort_values("bloom_index", ascending=False)

    ncols = len(MAIN_COLS)
    title_row(ws, 1, f"EXPIRING CONTRACTS 2026 — {len(exp)} PLAYERS  ·  Free transfer or cut-price acquisition opportunity", ncols,
              bg=AMBER_M, fg=WHITE, size=12, height=26)
    subtitle_row(ws, 2,
        "Contract expires Jun 2026 — can be pre-contracted from Jan 2026 for free or acquired at heavy discount  ·  Ranked by Lamberts Index",
        ncols, bg=AMBER_L, fg=AMBER_D)

    df_exp = build_df(exp)
    for col in df_exp.select_dtypes(include="float").columns:
        df_exp[col] = df_exp[col].round(2)
    write_table(ws, df_exp, "", start_row=2)


# ══════════════════════════════════════════════════════════════════════════════
# BUDGET PLANNER SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_budget_planner(ws, df):
    ws.title = "Budget Planner"
    ws.sheet_view.showGridLines = False

    BUDGET = 1_000_000

    def cell_w(r, c, v, bold=False, color=BLACK, bg=WHITE, align_h="left", fmt=None, merge_to=None):
        cc = ws.cell(r, c, v)
        cc.fill = fill(bg); cc.font = font(bold=bold, color=color, size=9)
        cc.alignment = align(align_h, "center")
        cc.border = border_bottom()
        if fmt: cc.number_format = fmt
        if merge_to: ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
        return cc

    title_row(ws, 1, "BUDGET PLANNER — FC Hradec Králové  ·  Budget Cap: €1,000,000", 15, bg=NAVY)
    subtitle_row(ws, 2,
        "ELITE VALUE + HIGH VALUE clear upgrades sorted by market value (cheapest first). Build your squad within budget.", 15)

    # Scenario summary box
    row = 3
    ws.row_dimensions[row].height = 20
    cell_w(row, 1, "BUDGET SUMMARY", bold=True, color=WHITE, bg=NAVY, merge_to=4)
    cell_w(row, 5, "Total Budget", bold=True, color=WHITE, bg=NAVY)
    cell_w(row, 6, BUDGET, bold=True, color=WHITE, bg=NAVY, fmt='€#,##0')
    row += 1

    # Instructions
    cell_w(row, 1, "Sort by Market Value to see cheapest options first. Running Total turns RED when budget is exceeded.", color=BLUE, bg=BLUE_LIGHT, merge_to=15)
    ws.row_dimensions[row].height = 14; row += 1

    # Filter to elite/high clear upgrades, sort by price
    budget_df = df[
        (df["upgrade_flag"] == "CLEAR UPGRADE") &
        (df["value_tier"].isin(["ELITE VALUE", "HIGH VALUE"]))
    ].sort_values("Market value", ascending=True, na_position="last")

    hdr_row = row
    hdr_vals = ["#", "Player", "Pos", "Team", "League", "Age", "Contract",
                "Mkt Val (€)", "Model Val (€)", "Val Ratio", "SQS", "Lamberts LI",
                "vs HK", "Tier", "Running Total (€)"]
    for ci2, v in enumerate(hdr_vals, 1):
        c = ws.cell(row, ci2, v)
        c.fill = fill(NAVY); c.font = font(bold=True, color=WHITE, size=9)
        c.alignment = align("center"); c.border = border_all("334155")
    ws.row_dimensions[row].height = 18
    row += 1

    running = 0
    for ri, (_, r2) in enumerate(budget_df.iterrows()):
        mv = int(r2.get("Market value", 0) or 0)
        running += mv
        over = running > BUDGET
        bg = RED_L if over else GREEN_L if ri % 2 == 0 else "E8FEF0"
        ws.row_dimensions[row].height = 15

        vals = [
            ri + 1,
            str(r2["Player"]),
            str(r2["pos_group"]),
            str(r2["Team"]),
            str(r2["league"]),
            int(r2.get("Age", 0) or 0),
            str(r2.get("Contract expires",""))[:10],
            mv,
            int(r2.get("model_value", 0) or 0),
            round(r2.get("value_ratio", 0) or 0, 1),
            round(r2.get("sqs_rank", 0) or 0, 1),
            round(r2.get("bloom_index", 0) or 0, 1),
            round(r2.get("vs_hradec_gap", 0) or 0, 1),
            str(r2.get("value_tier","")),
            running,
        ]
        fmts = [None, None, None, None, None, None, None,
                '€#,##0', '€#,##0', None, None, None, None, None, '€#,##0']
        aligns = ["center","left","center","left","left","center","center",
                  "right","right","center","center","center","center","center","right"]

        for ci2, (v2, fmt2, aln) in enumerate(zip(vals, fmts, aligns), 1):
            c = ws.cell(row, ci2, v2)
            c.fill = fill(bg); c.font = font(size=9); c.alignment = align(aln, "center")
            c.border = border_bottom()
            if fmt2: c.number_format = fmt2

        # Bold name
        ws.cell(row, 2).font = font(bold=True, color=RED_D if over else NAVY, size=9)
        # Colour bloom
        bi_v = vals[11]
        col_bi = GREEN_M if bi_v >= 20 else BLUE if bi_v >= 10 else GREY_D
        ws.cell(row, 12).font = font(bold=True, color=col_bi, size=9)
        # Tier colour
        tier_v = vals[13]
        t_bg2, t_fg2 = TIER_FILLS.get(tier_v, (GREY_BG, GREY_MID))
        ws.cell(row, 14).fill = fill(t_bg2); ws.cell(row, 14).font = font(bold=True, color=t_fg2, size=9)
        # Running total highlight
        ws.cell(row, 15).font = font(bold=True, color=RED_D if over else GREEN_D, size=9)
        if over:
            ws.cell(row, 15).fill = fill(RED_L)

        row += 1

    # Column widths
    widths = [5, 22, 6, 20, 13, 6, 11, 12, 13, 10, 9, 10, 8, 14, 14]
    for ci2, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci2)].width = w

    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(15)}{hdr_row}"
    ws.freeze_panes = f"B{hdr_row+1}"


# ══════════════════════════════════════════════════════════════════════════════
# SQUAD SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_squad_sheet(ws, squad_df, df_targets):
    ws.title = "Squad"
    ws.sheet_view.showGridLines = False

    squad_df = squad_df.copy()
    squad_df["pos_group"] = squad_df["position"].map(IMPECT_MAP).fillna("CM")
    squad_df["Quality %"] = (squad_df["IMPECT_SCORE_PACKING_pct"] * 100).round(1)

    out_cols = {
        "commonname":                     "Player",
        "position":                        "Position (Impect)",
        "pos_group":                       "Pos",
        "age":                             "Age",
        "playDuration":                    "Minutes",
        "Quality %":                       "Quality %",
        "IMPECT_SCORE_PACKING_pct":               "Overall Pctile",
        "OFFENSIVE_IMPECT_SCORE_PACKING_pct":     "Offensive Pctile",
        "DEFENSIVE_IMPECT_SCORE_PACKING_pct":     "Defensive Pctile",
        "PROGRESSION_SCORE_PACKING_pct":          "Progression Pctile",
    }
    existing = {k: v for k, v in out_cols.items() if k in squad_df.columns}
    df_sq = squad_df[list(existing.keys())].rename(columns=existing).copy()
    for col in df_sq.select_dtypes(include="float").columns:
        df_sq[col] = df_sq[col].round(3)
    df_sq = df_sq.sort_values("Minutes", ascending=False)

    ncols = len(df_sq.columns) + 2
    title_row(ws, 1, "FC HRADEC KRÁLOVÉ — CURRENT SQUAD QUALITY (Impect API)", ncols, bg=NAVY)
    subtitle_row(ws, 2, "Czech top-flight league percentile ranks  ·  Higher = better  ·  Basis for vs Hradec gap calculation", ncols)

    hrow = 3
    ws.append(list(df_sq.columns) + ["Gap Avail", "Best Target"])
    style_header(ws, hrow)
    ws.row_dimensions[hrow].height = 18

    # Pre-build best targets per position
    best_by_pos = {}
    for pos in POSITION_ORDER:
        sub = df_targets[df_targets["pos_group"] == pos].sort_values("bloom_index", ascending=False)
        if len(sub): best_by_pos[pos] = sub.iloc[0]

    for r_offset, row in enumerate(df_sq.itertuples(index=False)):
        r = hrow + 1 + r_offset
        ws.append(list(row))
        ws.row_dimensions[r].height = 15
        bg = STRIPE if r_offset % 2 == 0 else WHITE
        for cell in ws[r]:
            cell.fill = fill(bg); cell.font = font(size=9)
            cell.alignment = align("left", "center"); cell.border = border_bottom()

        q_col = list(df_sq.columns).index("Quality %") + 1
        q_val = ws.cell(r, q_col).value
        try:
            q_val = float(q_val)
            col = GREEN_M if q_val >= 70 else AMBER_M if q_val >= 40 else RED_M
            ws.cell(r, q_col).font = font(bold=True, color=col, size=10)
            ws.cell(r, q_col).alignment = align("center")
        except: pass

        # Percentile colour scale formatting per cell
        pct_cols = ["Overall Pctile", "Offensive Pctile", "Defensive Pctile", "Progression Pctile"]
        for pc in pct_cols:
            if pc in list(df_sq.columns):
                ci2 = list(df_sq.columns).index(pc) + 1
                try:
                    v = float(ws.cell(r, ci2).value or 0)
                    col = GREEN_M if v >= 0.7 else AMBER_M if v >= 0.4 else RED_M
                    ws.cell(r, ci2).font = font(bold=True, color=col, size=9)
                    ws.cell(r, ci2).alignment = align("center")
                except: pass

        # Gap available and best target
        pos_val = row[list(df_sq.columns).index("Pos")]
        if pos_val in best_by_pos:
            best = best_by_pos[pos_val]
            gap_ci = len(df_sq.columns) + 1
            best_ci = len(df_sq.columns) + 2
            gap_v = round(float(best.get("vs_hradec_gap", 0) or 0), 1)
            ws.cell(r, gap_ci, gap_v).fill = fill(GREEN_L if gap_v > 10 else bg)
            ws.cell(r, gap_ci).font = font(bold=True, color=GREEN_M if gap_v > 10 else GREY_D, size=9)
            ws.cell(r, gap_ci).alignment = align("center")
            ws.cell(r, best_ci, f"{best['Player']} (BI {round(float(best['bloom_index'] or 0),1)})").fill = fill(bg)
            ws.cell(r, best_ci).font = font(size=9, color=GREEN_D)
            ws.cell(r, best_ci).alignment = align("left")

    col_widths = {"A": 22, "B": 25, "C": 8, "D": 6, "E": 10, "F": 12,
                  "G": 14, "H": 16, "I": 16, "J": 18, "K": 12, "L": 30}
    for cl, w in col_widths.items():
        ws.column_dimensions[cl].width = w

    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(ncols)}{hrow}"
    ws.freeze_panes = "A4"

    # CF on percentile columns
    for col_name in ["Overall Pctile", "Offensive Pctile", "Defensive Pctile", "Progression Pctile"]:
        if col_name in list(df_sq.columns):
            cl = get_column_letter(list(df_sq.columns).index(col_name) + 1)
            ws.conditional_formatting.add(f"{cl}4:{cl}{3+len(df_sq)}", ColorScaleRule(
                start_type="num", start_value=0,   start_color="C0392B",
                mid_type="num",   mid_value=0.5,   mid_color="FFFFFF",
                end_type="num",   end_value=1.0,   end_color="1E8449",
            ))

    # ── Gap analysis section
    gap_start = hrow + len(df_sq) + 3
    ws.cell(gap_start, 1, "POSITION COVERAGE GAP ANALYSIS")
    ws.merge_cells(start_row=gap_start, start_column=1, end_row=gap_start, end_column=ncols)
    c = ws.cell(gap_start, 1)
    c.fill = fill(NAVY); c.font = font(bold=True, color=WHITE, size=11)
    c.alignment = align("left", "center", indent=1)
    ws.row_dimensions[gap_start].height = 22
    gap_start += 1

    hdr_g = ["Position", "Clear Upgrades Available", "Best BI", "Best Candidate",
             "Best Candidate Team", "Priority Level"]
    for ci2, v in enumerate(hdr_g, 1):
        c = ws.cell(gap_start, ci2, v)
        c.fill = fill(BLUE_LIGHT); c.font = font(bold=True, color=NAVY, size=9)
        c.alignment = align("center"); c.border = border_all("93C5FD")
    ws.row_dimensions[gap_start].height = 16
    gap_start += 1

    for pi, pos in enumerate(POSITION_ORDER):
        pdf = df_targets[df_targets["pos_group"] == pos]
        clear_n = (pdf["upgrade_flag"] == "CLEAR UPGRADE").sum()
        best = pdf.nlargest(1, "bloom_index").iloc[0] if len(pdf) else None
        pri = "🔴 HIGH PRIORITY" if clear_n >= 5 else "🟡 MEDIUM" if clear_n >= 2 else "🟢 COVERED"
        pri_bg = RED_L if "HIGH" in pri else AMBER_L if "MEDIUM" in pri else GREEN_L
        bg = STRIPE if pi % 2 == 0 else WHITE
        ws.row_dimensions[gap_start].height = 15

        ws.cell(gap_start, 1, f"{pos} — {POSITION_LABELS[pos]}").fill = fill(bg)
        ws.cell(gap_start, 1).font = font(bold=True, color=NAVY, size=9)
        ws.cell(gap_start, 1).alignment = align("left")

        ws.cell(gap_start, 2, clear_n).fill = fill(bg)
        ws.cell(gap_start, 2).font = font(bold=True, color=GREEN_M if clear_n > 3 else AMBER_M, size=9)
        ws.cell(gap_start, 2).alignment = align("center")

        if best is not None:
            bi_v = round(float(best.get("bloom_index", 0) or 0), 1)
            ws.cell(gap_start, 3, bi_v).fill = fill(bg)
            ws.cell(gap_start, 3).font = font(bold=True, color=GREEN_M if bi_v >= 20 else BLUE, size=9)
            ws.cell(gap_start, 3).alignment = align("center")
            ws.cell(gap_start, 4, str(best["Player"])).fill = fill(bg)
            ws.cell(gap_start, 4).font = font(bold=True, color=GREEN_D, size=9)
            ws.cell(gap_start, 4).alignment = align("left")
            ws.cell(gap_start, 5, str(best["Team"])).fill = fill(bg); ws.cell(gap_start, 5).font = font(size=9)
            ws.cell(gap_start, 5).alignment = align("left")

        ws.cell(gap_start, 6, pri).fill = fill(pri_bg)
        ws.cell(gap_start, 6).font = font(bold=True, size=9)
        ws.cell(gap_start, 6).alignment = align("center")
        gap_start += 1

    ws.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def run():
    print("Building FC Hradec Králové extensive workbook ...")

    df_raw   = pd.read_excel(DATA_FILE, sheet_name="All Targets (ranked)")
    squad_df = pd.read_excel(SQUAD_FILE)

    # Numeric coercion
    num_cols = ["Market value", "Age", "bloom_index", "sqs_rank",
                "Minutes played", "vs_hradec_gap", "model_value", "value_ratio"]
    for col in num_cols:
        if col in df_raw.columns:
            df_raw[col] = pd.to_numeric(df_raw[col], errors="coerce")

    # Derived columns
    df_raw["value_ratio"] = (df_raw["model_value"] / df_raw["Market value"].replace(0, np.nan)).fillna(0).round(2)
    df_raw["contract_flag"] = df_raw["Contract expires"].fillna("").astype(str).apply(
        lambda x: "2026" if "2026" in x else ("2027" if "2027" in x else ""))

    # Normalise value_tier
    def norm_tier(x):
        x = str(x).upper()
        if "ELITE" in x: return "ELITE VALUE"
        if "HIGH" in x:  return "HIGH VALUE"
        if x == "VALUE": return "VALUE"
        if "FAIR" in x:  return "FAIR PRICE"
        if "OVER" in x:  return "OVERVALUED"
        return str(x)

    if "value_tier" in df_raw.columns:
        df_raw["value_tier"] = df_raw["value_tier"].apply(norm_tier)
    else:
        df_raw["value_tier"] = df_raw["bloom_index"].apply(
            lambda b: "ELITE VALUE" if b >= 30 else "HIGH VALUE" if b >= 20 else
                      "VALUE" if b >= 10 else "FAIR PRICE" if b >= 0 else "OVERVALUED")

    # Format val ratio as string with × for display
    df_raw["value_ratio_str"] = df_raw["value_ratio"].apply(lambda x: f"{x:.1f}×" if x > 0 else "—")

    squad_df["playDuration"] = pd.to_numeric(squad_df["playDuration"], errors="coerce").fillna(0)
    squad_df["IMPECT_SCORE_PACKING_pct"] = pd.to_numeric(
        squad_df["IMPECT_SCORE_PACKING_pct"], errors="coerce").fillna(0)

    # Add value_ratio_str to MAIN_COLS mapping for build_df
    df_raw["val_ratio_display"] = df_raw["value_ratio_str"]

    # Patch MAIN_COLS to use string val ratio
    global MAIN_COLS
    MAIN_COLS_PATCHED = []
    for item in MAIN_COLS:
        if item[0] == "value_ratio":
            MAIN_COLS_PATCHED.append(("val_ratio_display", "Val Ratio", 10))
        else:
            MAIN_COLS_PATCHED.append(item)
    MAIN_COLS = MAIN_COLS_PATCHED

    wb = Workbook()

    # ── README
    print("  README ...")
    ws_readme = wb.active
    build_readme(ws_readme)

    # ── Priority List
    print("  Priority List ...")
    ws_pri = wb.create_sheet("Priority List")
    ws_pri.sheet_view.showGridLines = False
    priority = df_raw[df_raw["upgrade_flag"] == "CLEAR UPGRADE"].sort_values(
        "bloom_index", ascending=False, na_position="last")
    df_pri = build_df(priority)
    for col in df_pri.select_dtypes(include="float").columns:
        df_pri[col] = df_pri[col].round(2)
    write_table(ws_pri, df_pri,
                f"PRIORITY LIST — All {len(df_pri)} Clear Upgrades — Ranked by Lamberts Index",
                "Jamestown Analytics  ·  All are clear statistical upgrades on current Hradec starters  ·  Budget ≤ €1M")

    # ── Elite Picks
    print("  Elite Picks ...")
    ws_elite = wb.create_sheet("Elite Picks")
    ws_elite.sheet_view.showGridLines = False
    elite = df_raw[df_raw["value_tier"] == "ELITE VALUE"].sort_values(
        "bloom_index", ascending=False, na_position="last")
    df_elite = build_df(elite)
    for col in df_elite.select_dtypes(include="float").columns:
        df_elite[col] = df_elite[col].round(2)
    write_table(ws_elite, df_elite,
                f"ELITE VALUE — {len(df_elite)} Players with Lamberts Index ≥ 30",
                "Strongest buy signals  ·  Statistical output is 30+ percentile ranks above what the market charges  ·  Sorted by Lamberts Index",
                table_style="TableStyleMedium7")

    # ── Lamberts Analysis
    print("  Lamberts Analysis ...")
    ws_bloom = wb.create_sheet("Lamberts Analysis")
    build_bloom_analysis(ws_bloom, df_raw)

    # ── Top 5 per Position
    print("  Top 5 per Position ...")
    ws_top5 = wb.create_sheet("Top 5 per Pos")
    build_top5_sheet(ws_top5, df_raw)

    # ── All Targets
    print("  All Targets ...")
    ws_all = wb.create_sheet("All Targets")
    ws_all.sheet_view.showGridLines = False
    all_t = df_raw.sort_values("bloom_index", ascending=False, na_position="last")
    df_all = build_df(all_t)
    for col in df_all.select_dtypes(include="float").columns:
        df_all[col] = df_all[col].round(2)
    write_table(ws_all, df_all,
                f"ALL TARGETS — {len(df_all)} Candidates  ·  CZ II + Slovak leagues",
                "Use column filters to narrow by position, league, age, status, tier  ·  Ranked by Lamberts Index")

    # ── Per-position sheets
    for pg in POSITION_ORDER:
        print(f"  {pg} ...")
        ws_pg = wb.create_sheet(pg)
        ws_pg.sheet_view.showGridLines = False
        sub = df_raw[df_raw["pos_group"] == pg].sort_values(
            "bloom_index", ascending=False, na_position="last")
        if len(sub) == 0: continue
        df_pg = build_df(sub)
        for col in df_pg.select_dtypes(include="float").columns:
            df_pg[col] = df_pg[col].round(2)
        write_table(ws_pg, df_pg,
                    f"{POSITION_LABELS[pg].upper()} TARGETS  ·  {len(df_pg)} candidates",
                    f"Sorted by Lamberts Index  ·  All {pg} candidates in recruitment universe")

    # ── Expiring 2026
    print("  Expiring 2026 ...")
    ws_exp = wb.create_sheet("Expiring 2026")
    build_expiring(ws_exp, df_raw)

    # ── Budget Planner
    print("  Budget Planner ...")
    ws_budget = wb.create_sheet("Budget Planner")
    build_budget_planner(ws_budget, df_raw)

    # ── Squad
    print("  Squad ...")
    ws_squad = wb.create_sheet("Squad")
    build_squad_sheet(ws_squad, squad_df, df_raw)

    # Tab colours
    tab_colors = {
        "README":         "1E3A8A",
        "Priority List":  "166534",
        "Elite Picks":    "14532D",
        "Lamberts Analysis": "1A56DB",
        "Top 5 per Pos":  "2563EB",
        "All Targets":    "374151",
        "GK":  "0369A1", "CB": "0369A1", "FB": "0369A1",
        "DM":  "0369A1", "CM": "0369A1", "W":  "0369A1", "FW": "0369A1",
        "Expiring 2026":  "92400E",
        "Budget Planner": "6D28D9",
        "Squad":          "D97706",
    }
    for sn, color in tab_colors.items():
        if sn in wb.sheetnames:
            wb[sn].sheet_properties.tabColor = color

    wb.save(OUT_FILE)
    total_sheets = len(wb.sheetnames)
    print(f"\n✓ Done → {OUT_FILE}")
    print(f"  {total_sheets} sheets: {' · '.join(wb.sheetnames)}")


if __name__ == "__main__":
    run()
